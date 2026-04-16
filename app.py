# ============================================================
#  app.py — IT Procurement Intelligence Dashboard
#  PwC Brand | Source Sans Pro
#  - Master Catalog upload tab
#  - AI auto-builds dashboard from uploaded catalog
#  - Fixed expander arrow overlap
#  - Subcategorized service selection
#  - Price score in comparison table
#  - AI-powered vendor & service intelligence
# ============================================================

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from collections import defaultdict
import openpyxl
import os, re, io, zipfile, difflib

try:
    import requests
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False

# ════════════════════════════════════════════════════════════
# PAGE CONFIG
# ════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="IT Procurement Intelligence",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ════════════════════════════════════════════════════════════
# CSS
# ════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Source+Sans+Pro:wght@300;400;600;700&display=swap');

html, body, [class*="css"], div, p, span, td, th,
label, button, .stMarkdown {
    font-family: 'Source Sans Pro','Helvetica Neue',
                 Arial, sans-serif !important;
}
.main .block-container {
    background-color : #F3F3F3 !important;
    padding-top      : 1.5rem;
    max-width        : 100% !important;
    padding-left     : 2rem !important;
    padding-right    : 2rem !important;
}
#MainMenu {visibility:hidden;}
footer    {visibility:hidden;}
header    {visibility:hidden;}
[data-testid="collapsedControl"] { display:none !important; }

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background-color : #2D2D2D !important;
    border-right     : 3px solid #D04A02;
    min-width        : 310px !important;
    max-width        : 310px !important;
}
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div {
    color: #F0F0F0 !important;
    font-family: 'Source Sans Pro', sans-serif !important;
}
section[data-testid="stSidebar"] div[data-baseweb="select"] {
    background-color:#FFFFFF !important;
    border-radius:2px !important;
    border:1px solid #999 !important;
}
section[data-testid="stSidebar"] div[data-baseweb="select"] * {
    color:#2D2D2D !important;
}
section[data-testid="stSidebar"] div[data-baseweb="input"] {
    background-color:#FFFFFF !important;
    border-radius:2px !important;
}
section[data-testid="stSidebar"] div[data-baseweb="input"] input {
    color:#2D2D2D !important;
}
section[data-testid="stSidebar"] span[data-baseweb="tag"] {
    background-color:#D04A02 !important;
    border-radius:2px !important;
}
section[data-testid="stSidebar"] span[data-baseweb="tag"] span {
    color:white !important;
}

/* ── KPI ── */
.kpi-box {
    border-radius:4px; padding:18px 10px;
    text-align:center; color:white;
    border-left:5px solid rgba(255,255,255,0.25);
}
.kpi-value {
    font-size:2.1em; font-weight:700;
    margin:0; line-height:1.1; letter-spacing:-0.5px;
}
.kpi-label {
    font-size:0.78em; font-weight:700; opacity:0.9;
    margin-top:5px; letter-spacing:0.8px;
    text-transform:uppercase;
}

/* ── Tabs ── */
button[data-baseweb="tab"] {
    font-weight:600 !important;
    font-size:0.92em !important;
    color:#7D7D7D !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color:#D04A02 !important;
    border-bottom:3px solid #D04A02 !important;
}

/* ── FIX: Remove arrow overlap from expander ── */
div[data-testid="stExpander"] details > summary {
    list-style   : none !important;
    padding-left : 12px !important;
}
div[data-testid="stExpander"] details > summary::before,
div[data-testid="stExpander"] details > summary::after,
div[data-testid="stExpander"] details > summary::-webkit-details-marker,
div[data-testid="stExpander"] details > summary::marker {
    display : none !important;
    content : "" !important;
    width   : 0 !important;
}
div[data-testid="stExpander"] details summary p {
    font-weight  : 700;
    font-size    : 0.94em;
    color        : #2D2D2D !important;
    padding-left : 0 !important;
    margin-left  : 0 !important;
}
div[data-testid="stExpander"] details {
    border        : 1px solid #ddd;
    border-radius : 4px;
    margin-bottom : 10px;
    padding       : 2px 0;
}

/* ── Comparison table ── */
.comp-table {
    width:100%; border-collapse:collapse;
    table-layout:fixed; font-size:0.83em;
    border:1px solid #e0e0e0;
}
.comp-table thead tr { background:#2D2D2D; }
.comp-table thead th {
    padding:10px 10px; text-align:left;
    font-weight:700; font-size:0.80em;
    letter-spacing:0.4px; text-transform:uppercase;
    color:white !important; border:none;
    word-break:break-word;
}
.comp-table tbody tr:nth-child(even) { background:#F3F3F3; }
.comp-table tbody tr:hover           { background:#FCE8DC; }
.comp-table tbody td {
    padding:8px 10px; border-bottom:1px solid #e8e8e8;
    vertical-align:middle; word-break:break-word;
    font-size:0.82em; color:#2D2D2D;
}
.comp-table th:nth-child(1),
.comp-table td:nth-child(1) { width:13%; }
.comp-table th:nth-child(2),
.comp-table td:nth-child(2) { width:11%; }
.comp-table th:nth-child(3),
.comp-table td:nth-child(3) { width:22%; }
.comp-table th:nth-child(4),
.comp-table td:nth-child(4) { width:11%; }
.comp-table th:nth-child(5),
.comp-table td:nth-child(5) { width:11%; }
.comp-table th:nth-child(6),
.comp-table td:nth-child(6) { width:12%; }
.comp-table th:nth-child(7),
.comp-table td:nth-child(7) { width:10%; }
.comp-table th:nth-child(8),
.comp-table td:nth-child(8) { width:10%; }

/* ── Vendor badge ── */
.vendor-badge {
    display:inline-block; padding:3px 8px;
    border-radius:2px; color:white;
    font-size:0.78em; font-weight:700;
    white-space:nowrap; overflow:hidden;
    text-overflow:ellipsis; max-width:100%;
    box-sizing:border-box;
}

/* ── Score cards ── */
.score-card {
    border-radius:4px; padding:14px 16px;
    margin-bottom:10px;
    border-left:5px solid #D04A02;
}
.score-high   { background:#FFF3F0; border-color:#E0301E; }
.score-medium { background:#FFF8E1; border-color:#FFB600; }
.score-low    { background:#F0FFF4; border-color:#22992E; }
.score-num {
    font-size:2.2em; font-weight:800;
    line-height:1; letter-spacing:-1px;
}

/* ── AI box ── */
.ai-box {
    background:#F8F0FF;
    border-left:5px solid #6E2585;
    border-radius:4px;
    padding:14px 18px;
    margin:10px 0;
}
.ai-box-title {
    font-size:0.75em; font-weight:700;
    letter-spacing:1px; text-transform:uppercase;
    color:#6E2585; margin-bottom:6px;
}

/* ── Upload catalog box ── */
.upload-catalog-box {
    background        : white;
    border            : 2px dashed #D04A02;
    border-radius     : 8px;
    padding           : 30px;
    text-align        : center;
    margin-bottom     : 20px;
}
.catalog-step {
    background        : white;
    border-left       : 4px solid #D04A02;
    border-radius     : 4px;
    padding           : 16px 20px;
    margin-bottom     : 12px;
    box-shadow        : 0 1px 4px rgba(0,0,0,0.06);
}
.catalog-step-num {
    font-size         : 0.72em;
    font-weight       : 700;
    letter-spacing    : 1px;
    text-transform    : uppercase;
    color             : #D04A02;
    margin-bottom     : 4px;
}
.insight-card {
    background        : #F8F8F8;
    border-radius     : 4px;
    padding           : 14px 16px;
    margin-bottom     : 10px;
    border            : 1px solid #e0e0e0;
}
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# PwC COLOURS
# ════════════════════════════════════════════════════════════
COLORS = [
    "#D04A02","#295477","#299D8F",
    "#FFB600","#22992E","#E0301E",
    "#EB8C00","#6E2585","#8C8C8C","#004F9F",
]

def get_color(i):
    return COLORS[i % len(COLORS)]

CFONT = dict(
    family="Source Sans Pro, Helvetica Neue, Arial",
    size=11, color="#2D2D2D")
CBG = "#F3F3F3"


# ════════════════════════════════════════════════════════════
# PRICE EXTRACTION
# ════════════════════════════════════════════════════════════
PRICE_RE = re.compile(
    r"""
    (?:USD|EUR|GBP|SGD|MYR|AUD|CAD)\s?\d{1,3}(?:[,]\d{3})*(?:\.\d{1,2})?
    |(?:[\$\€\£]\s?)\d{1,3}(?:[,\s]\d{3})*(?:\.\d{1,2})?
    |\d{1,3}(?:[,]\d{3})+(?:\.\d{1,2})?
    """,
    re.VERBOSE | re.IGNORECASE,
)
TOTAL_KW = [
    "grand total","total amount","total price","amount due",
    "net total","total cost","total value","quote total",
    "subtotal","estimated total","total",
]


def _parse_num(s):
    try:
        return float(re.sub(r"[^\d.]", "", str(s)) or "0")
    except Exception:
        return 0.0


def _fmt(val):
    try:
        v = float(re.sub(r"[^\d.]", "", str(val)) or "0")
        if v <= 0:
            return "—"
        return "${:,.2f}".format(v)
    except Exception:
        return str(val)


def _best_price(text):
    tl = text.lower()
    for kw in TOTAL_KW:
        idx = tl.find(kw)
        if idx == -1:
            continue
        snippet = text[max(0, idx - 20): idx + 300]
        hits    = PRICE_RE.findall(snippet)
        valid   = [h.strip() for h in hits if _parse_num(h) >= 50]
        if valid:
            return max(valid, key=_parse_num)
    all_hits = PRICE_RE.findall(text)
    valid    = [h.strip() for h in all_hits if _parse_num(h) >= 100]
    if valid:
        return max(valid, key=_parse_num)
    return ""


def _text_from_bytes(content, ext):
    text = ""
    ext  = ext.lower().strip(".")
    try:
        if ext == "pdf":
            if not PDF_OK:
                return ""
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text += t + "\n"
        elif ext in ("xlsx", "xls"):
            wb = openpyxl.load_workbook(
                io.BytesIO(content), data_only=True,
                read_only=True)
            rows_text = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    rs = "  ".join(
                        str(c) for c in row if c is not None)
                    if rs.strip():
                        rows_text.append(rs)
            text = "\n".join(rows_text)
            wb.close()
        elif ext == "docx":
            with zipfile.ZipFile(io.BytesIO(content)) as z:
                if "word/document.xml" in z.namelist():
                    xml  = z.read("word/document.xml").decode(
                        "utf-8", errors="ignore")
                    text = re.sub(r"<[^>]+>", " ", xml)
                    text = re.sub(r"\s{2,}", "\n", text)
    except Exception:
        pass
    return text


def extract_price_from_bytes(content, ext):
    text  = _text_from_bytes(content, ext)
    price = _best_price(text)
    return {
        "price"    : price,
        "price_num": _parse_num(price) if price else 0.0,
        "text"     : text[:5000],
    }


def extract_price_from_url(url):
    if not REQUESTS_OK:
        return {"price": "", "price_num": 0.0, "text": ""}
    try:
        r   = requests.get(url, timeout=20)
        ext = url.split("?")[0].rsplit(".", 1)[-1].lower()
        return extract_price_from_bytes(r.content, ext)
    except Exception:
        return {"price": "", "price_num": 0.0, "text": ""}


# ════════════════════════════════════════════════════════════
# SIMILARITY
# ════════════════════════════════════════════════════════════
def similarity_score(text_a, text_b):
    def clean(t):
        t      = re.sub(r"[^\w\s]", " ", t.lower())
        tokens = [w for w in t.split() if len(w) > 2]
        return " ".join(sorted(set(tokens)))
    a = clean(text_a[:3000])
    b = clean(text_b[:3000])
    if not a or not b:
        return 0
    return round(
        difflib.SequenceMatcher(None, a, b).ratio() * 100, 1)


def service_similarity(sa, sb):
    if not sa or not sb:
        return 0.0
    a = set(s.lower().strip() for s in sa)
    b = set(s.lower().strip() for s in sb)
    if not a or not b:
        return 0.0
    return round(len(a & b) / len(a | b) * 100, 1)


def price_score(new_price, hist_prices):
    valid = [p for p in hist_prices if p > 0]
    if not valid or new_price <= 0:
        return None, "No comparison data"
    mn  = min(valid)
    mx  = max(valid)
    avg = sum(valid) / len(valid)
    if mx == mn:
        return 50, "Same as historical average"
    score = round((1 - (new_price - mn) / (mx - mn)) * 100, 1)
    score = max(0, min(100, score))
    pct   = round((new_price - avg) / avg * 100, 1)
    if new_price < avg:
        label = "{}% BELOW average — COMPETITIVE".format(abs(pct))
    elif new_price > avg:
        label = "{}% ABOVE average — REVIEW NEEDED".format(
            abs(pct))
    else:
        label = "Matches historical average"
    return score, label


def score_color(s):
    if s is None:
        return "#8C8C8C"
    if s >= 70:
        return "#22992E"
    if s >= 40:
        return "#FFB600"
    return "#E0301E"


def score_css(s):
    if s is None:
        return "score-medium"
    if s >= 70:
        return "score-low"
    if s >= 40:
        return "score-medium"
    return "score-high"


# ════════════════════════════════════════════════════════════
# AI INSIGHTS
# ════════════════════════════════════════════════════════════
def ai_vendor_insight(vendor, services, prices):
    n_svc = len(services)
    n_prc = len([p for p in prices if p > 0])
    lines = [
        "**{}** offers **{}** service(s).".format(vendor, n_svc)
    ]
    if n_prc > 0:
        avg_p = sum(prices) / n_prc
        lines.append(
            "Average quoted price: **{}**.".format(_fmt(avg_p)))
    if n_svc >= 5:
        lines.append(
            "Broad coverage — suitable for consolidated "
            "procurement.")
    elif n_svc >= 2:
        lines.append(
            "Moderate coverage — consider combining vendors.")
    else:
        lines.append(
            "Limited coverage — best as specialist supplier.")
    return " ".join(lines)


def ai_price_insight(new_price, hist_prices, vendor_prices):
    valid = [p for p in hist_prices if p > 0]
    if not valid or new_price <= 0:
        return "Insufficient data for price analysis."
    avg  = sum(valid) / len(valid)
    mn   = min(valid)
    mx   = max(valid)
    pct  = round((new_price - avg) / avg * 100, 1)
    lines = []
    if new_price <= mn:
        lines.append(
            "This quote is the **lowest price** seen — "
            "excellent value.")
    elif new_price >= mx:
        lines.append(
            "This quote is **above all historical prices** — "
            "negotiate strongly.")
    elif pct > 15:
        lines.append(
            "Quote is **{}% above** average. "
            "Request a revised quote.".format(abs(pct)))
    elif pct < -15:
        lines.append(
            "Quote is **{}% below** average — "
            "very competitive.".format(abs(pct)))
    else:
        lines.append(
            "Quote is **within normal range** "
            "({}% vs average).".format(pct))
    if vendor_prices:
        best_v = min(vendor_prices, key=vendor_prices.get)
        lines.append(
            "**{}** has historically offered the lowest "
            "prices.".format(best_v))
    return " ".join(lines)


def ai_service_summary(services_by_vendor):
    if not services_by_vendor:
        return "No vendor data available."
    best   = max(services_by_vendor,
                 key=lambda v: len(services_by_vendor[v]))
    n_best = len(services_by_vendor[best])
    total  = len(set(
        s for svcs in services_by_vendor.values()
        for s in svcs))
    lines  = [
        "**{}** covers the most services "
        "({} of {} total).".format(best, n_best, total)
    ]
    shared = [
        s for s in set(
            s for svcs in services_by_vendor.values()
            for s in svcs)
        if sum(1 for svcs in services_by_vendor.values()
               if s in svcs) > 1
    ]
    if shared:
        lines.append(
            "**{}** service(s) offered by multiple vendors — "
            "ideal for competitive benchmarking.".format(
                len(shared)))
    return " ".join(lines)


# ════════════════════════════════════════════════════════════
# AI CATALOG ANALYZER
# ════════════════════════════════════════════════════════════
def ai_analyze_catalog(df, df_exp):
    """
    Generates a comprehensive AI analysis of an uploaded
    master catalog. Returns dict of insight sections.
    """
    insights = {}

    # Basic stats
    n_vendors  = df["Vendor"].nunique()
    n_files    = df["File Name"].nunique()
    n_cats     = df["Category"].nunique()
    n_services = df_exp["Service"].nunique()

    insights["overview"] = (
        "Catalog contains **{} vendors**, **{} quote files**, "
        "**{} categories** and **{} unique services**. "
        .format(n_vendors, n_files, n_cats, n_services)
    )

    # Top vendor by services
    spv = (df_exp.groupby("Vendor")["Service"]
           .nunique().sort_values(ascending=False))
    if not spv.empty:
        top_v = spv.index[0]
        insights["top_vendor"] = (
            "**{}** leads with **{}** unique services — "
            "strongest overall coverage.".format(
                top_v, spv.iloc[0])
        )

    # Most competitive services
    shared = (df_exp.groupby("Service")["Vendor"]
              .nunique().sort_values(ascending=False))
    hot = shared[shared > 1]
    if not hot.empty:
        top_svc = hot.index[0]
        insights["competitive"] = (
            "**{}** is the most competitive service with "
            "**{}** vendors quoting — strong negotiating "
            "position available.".format(
                top_svc, hot.iloc[0])
        )

    # Category concentration
    cat_counts = (
        df.drop_duplicates(subset=["Category", "File Name"])
        .groupby("Category").size()
        .sort_values(ascending=False)
    )
    if not cat_counts.empty:
        top_cat = cat_counts.index[0]
        pct     = round(
            cat_counts.iloc[0] / cat_counts.sum() * 100, 1)
        insights["category"] = (
            "**{}** dominates with **{}%** of all quote files — "
            "consider diversifying procurement scope.".format(
                top_cat, pct)
        )

    # Pricing insight
    if "Quoted Price" in df.columns:
        prices = df["Quoted Price"].apply(_parse_num)
        prices = prices[prices > 0]
        if not prices.empty:
            insights["pricing"] = (
                "Quoted prices range from **{}** to **{}** "
                "with an average of **{}**.".format(
                    _fmt(prices.min()),
                    _fmt(prices.max()),
                    _fmt(prices.mean()))
            )

    # Recommendations
    recs = []
    if not hot.empty and len(hot) >= 3:
        recs.append(
            "Run competitive bids on the {} services "
            "offered by multiple vendors.".format(len(hot)))
    if n_vendors >= 3:
        recs.append(
            "Consider vendor consolidation — "
            "{} vendors may create management overhead.".format(
                n_vendors))
    spv_low = spv[spv == 1]
    if len(spv_low) > 0:
        recs.append(
            "{} vendors offer only 1 service — "
            "evaluate if specialist contracts are "
            "justified.".format(len(spv_low)))
    insights["recommendations"] = recs

    return insights


def ai_detect_columns(df_raw):
    """
    AI-powered column detection for uploaded catalog.
    Maps columns to standard names regardless of header labels.
    """
    col_map = {}
    for c in df_raw.columns:
        cl = str(c).lower().strip()
        # Category
        if any(k in cl for k in ["category", "type", "domain"]):
            if "Category" not in col_map:
                col_map["Category"] = c
        # Vendor
        elif any(k in cl for k in
                 ["vendor", "supplier", "company",
                  "provider", "partner"]):
            if "Vendor" not in col_map:
                col_map["Vendor"] = c
        # File Name
        elif any(k in cl for k in
                 ["file name", "filename", "document",
                  "file", "attachment"]):
            if "File Name" not in col_map:
                col_map["File Name"] = c
        # File Link
        elif any(k in cl for k in
                 ["link", "url", "hyperlink", "path"]):
            if "File Link" not in col_map:
                col_map["File Link"] = c
        # Comments / Services
        elif any(k in cl for k in
                 ["comment", "service", "description",
                  "scope", "remark", "note"]):
            if "Comments" not in col_map:
                col_map["Comments"] = c
        # Price
        elif any(k in cl for k in
                 ["price", "cost", "amount",
                  "value", "quote", "rate"]):
            if "Quoted Price" not in col_map:
                col_map["Quoted Price"] = c
    return col_map


def process_uploaded_catalog(file_bytes, filename):
    """
    Processes an uploaded Excel/CSV catalog file.
    Returns (df_master, df_exploded) or (None, None).
    """
    try:
        ext = filename.rsplit(".", 1)[-1].lower()

        # Read raw
        if ext in ("xlsx", "xls"):
            # Try to detect header row
            raw = pd.read_excel(
                io.BytesIO(file_bytes),
                engine="openpyxl",
                header=None)
        elif ext == "csv":
            raw = pd.read_csv(io.BytesIO(file_bytes), header=None)
        else:
            return None, None, "Unsupported file type."

        # Detect header row
        header_row = 0
        for i, row in raw.iterrows():
            vals = [str(v).strip().lower()
                    for v in row.values if pd.notna(v)]
            has_vendor = any(
                k in " ".join(vals)
                for k in ["vendor","supplier","company"])
            has_file   = any(
                k in " ".join(vals)
                for k in ["file","document","attachment"])
            has_cat    = any(
                k in " ".join(vals)
                for k in ["category","type","domain"])
            if (has_vendor or has_cat) and has_file:
                header_row = i
                break

        # Re-read with correct header
        if ext in ("xlsx", "xls"):
            df = pd.read_excel(
                io.BytesIO(file_bytes),
                engine="openpyxl",
                header=header_row)
        else:
            df = pd.read_csv(
                io.BytesIO(file_bytes),
                header=header_row)

        df = df.loc[:, df.columns.notna()]
        df.columns = [str(c).strip() for c in df.columns]
        df.dropna(how="all", inplace=True)

        # AI column detection
        col_map = ai_detect_columns(df)

        if len(col_map) < 2:
            return None, None, (
                "Could not detect required columns. "
                "Please ensure your file has Vendor, "
                "Category, and File Name columns.")

        df.rename(
            columns={v: k for k, v in col_map.items()},
            inplace=True)

        # Ensure required columns exist
        for req in ["Category", "Vendor", "File Name"]:
            if req not in df.columns:
                df[req] = ""

        if "Comments" not in df.columns:
            df["Comments"] = ""

        keep = ["Category", "Vendor", "File Name", "Comments"]
        for e in ["File Link", "Quoted Price"]:
            if e in df.columns:
                keep.append(e)

        df = df[[c for c in keep if c in df.columns]].copy()

        df = df[
            ~(df["Category"].astype(str).str.strip().isin(
                ["", "nan"]) &
              df["Vendor"].astype(str).str.strip().isin(
                  ["", "nan"]))
        ].copy()

        for col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()
        df.reset_index(drop=True, inplace=True)

        def parse_svc(v):
            if not v or str(v).strip() in ["", "nan"]:
                return ["(unspecified)"]
            parts = [s.strip()
                     for s in str(v).split("\n") if s.strip()]
            return parts or ["(unspecified)"]

        df["Services List"] = df["Comments"].apply(parse_svc)

        df_exp = df.explode("Services List").copy()
        df_exp.rename(
            columns={"Services List": "Service"}, inplace=True)
        df_exp["Service"] = df_exp["Service"].str.strip()
        df_exp = df_exp[
            ~df_exp["Service"].isin(
                ["", "(unspecified)", "nan"])
        ].reset_index(drop=True)

        return df, df_exp, None

    except Exception as e:
        return None, None, "Error: {}".format(e)


# ════════════════════════════════════════════════════════════
# SUBCATEGORIZE SERVICES
# ════════════════════════════════════════════════════════════
def group_services(services):
    groups = defaultdict(list)
    for svc in sorted(services):
        tokens = str(svc).strip().split()
        if not tokens:
            groups["Other"].append(svc)
            continue
        key = tokens[0]
        if len(key) <= 3 and len(tokens) > 1:
            key = tokens[0] + " " + tokens[1]
        key = key.strip()
        groups[key].append(svc)
    final = {}
    other = []
    for k, v in groups.items():
        if len(v) >= 2:
            final[k] = sorted(v)
        else:
            other.extend(v)
    if other:
        final["Other"] = sorted(other)
    return dict(sorted(final.items()))


# ════════════════════════════════════════════════════════════
# HYPERLINK EXTRACTION
# ════════════════════════════════════════════════════════════
@st.cache_data
def extract_hyperlinks(file_path):
    link_map = {}
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        fn_col = hr = None
        for row in ws.iter_rows():
            for cell in row:
                if (cell.value and
                        str(cell.value).strip().lower()
                        == "file name"):
                    fn_col = cell.column
                    hr     = cell.row
                    break
            if fn_col:
                break
        if fn_col:
            for row in ws.iter_rows(
                    min_row=hr + 1,
                    min_col=fn_col, max_col=fn_col):
                cell = row[0]
                if cell.value and cell.hyperlink:
                    link_map[str(cell.value).strip()] = \
                        str(cell.hyperlink.target).strip()
        wb.close()
    except Exception as e:
        st.warning("Hyperlink warning: {}".format(e))
    return link_map


# ════════════════════════════════════════════════════════════
# DATA LOADING — default file
# ════════════════════════════════════════════════════════════
@st.cache_data
def load_data():
    FILE_PATH = "Master Catalog.xlsx"
    if not os.path.exists(FILE_PATH):
        return None, None

    raw = pd.read_excel(
        FILE_PATH, engine="openpyxl", header=None)
    header_row = None
    for i, row in raw.iterrows():
        vals = [str(v).strip().lower()
                for v in row.values if pd.notna(v)]
        if (any("category" in v for v in vals) and
                any("file" in v for v in vals)):
            header_row = i
            break
    if header_row is None:
        return None, None

    df = pd.read_excel(
        FILE_PATH, engine="openpyxl", header=header_row)
    df = df.loc[:, df.columns.notna()]
    df.columns = [str(c).strip() for c in df.columns]
    df.dropna(how="all", inplace=True)

    col_map = {}
    for c in df.columns:
        cl = str(c).lower().strip()
        if cl == "category":
            col_map["Category"]     = c
        elif "vendor" in cl or "type" in cl:
            col_map["Vendor"]       = c
        elif cl == "file name":
            col_map["File Name"]    = c
        elif cl == "file link":
            col_map["File Link"]    = c
        elif cl == "file url":
            col_map["File URL"]     = c
        elif "comment" in cl:
            col_map["Comments"]     = c
        elif "quoted" in cl or "price" in cl:
            col_map["Quoted Price"] = c

    df.rename(
        columns={v: k for k, v in col_map.items()},
        inplace=True)

    keep = ["Category", "Vendor", "File Name", "Comments"]
    for e in ["File Link", "File URL", "Quoted Price"]:
        if e in df.columns:
            keep.append(e)
    df = df[[c for c in keep if c in df.columns]].copy()

    df = df[
        ~(df["Category"].astype(str).str.strip().isin(
            ["", "nan"]) &
          df["Vendor"].astype(str).str.strip().isin(
              ["", "nan"]))
    ].copy()

    for col in df.columns:
        df[col] = df[col].fillna("").astype(str).str.strip()
    df.reset_index(drop=True, inplace=True)

    hmap = extract_hyperlinks(FILE_PATH)
    df["Hyperlink"] = df["File Name"].map(hmap).fillna("")
    for fb in ["File Link", "File URL"]:
        if fb in df.columns:
            df["Hyperlink"] = df.apply(
                lambda r: r["Hyperlink"]
                if r["Hyperlink"] not in ["", "nan"]
                else r[fb], axis=1)

    def parse_svc(v):
        if not v or str(v).strip() in ["", "nan"]:
            return ["(unspecified)"]
        parts = [s.strip()
                 for s in str(v).split("\n") if s.strip()]
        return parts or ["(unspecified)"]

    df["Services List"] = df["Comments"].apply(parse_svc)

    df_exp = df.explode("Services List").copy()
    df_exp.rename(
        columns={"Services List": "Service"}, inplace=True)
    df_exp["Service"] = df_exp["Service"].str.strip()
    df_exp = df_exp[
        ~df_exp["Service"].isin(["", "(unspecified)", "nan"])
    ].reset_index(drop=True)

    return df, df_exp


# ════════════════════════════════════════════════════════════
# DETERMINE DATA SOURCE
# ════════════════════════════════════════════════════════════
# Check if user uploaded a catalog via session_state
if "uploaded_catalog_df" in st.session_state and \
        st.session_state["uploaded_catalog_df"] is not None:
    df_master  = st.session_state["uploaded_catalog_df"]
    df_exploded = st.session_state["uploaded_catalog_exp"]
    DATA_SOURCE = "uploaded"
else:
    df_master, df_exploded = load_data()
    DATA_SOURCE = "file"

# If no data at all, show only upload tab
NO_DATA = df_master is None or df_exploded is None

if not NO_DATA:
    vendor_color_map = {
        v: get_color(i)
        for i, v in enumerate(
            sorted(df_master["Vendor"].unique()))
    }
else:
    vendor_color_map = {}


# ════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════
def sb_label(txt):
    st.markdown(
        "<p style='color:#F0F0F0;font-weight:700;"
        "font-size:0.85em;margin:12px 0 4px;"
        "letter-spacing:0.5px;text-transform:uppercase'>"
        "{}</p>".format(txt),
        unsafe_allow_html=True)


def section_title(txt, caption=""):
    st.markdown(
        "<div style='font-size:0.78em;font-weight:700;"
        "letter-spacing:1px;text-transform:uppercase;"
        "color:#D04A02;margin-bottom:4px'>{}</div>".format(txt),
        unsafe_allow_html=True)
    if caption:
        st.caption(caption)


def vendor_pill(v, color):
    return (
        "<span class='vendor-badge' "
        "style='background:{}'>{}</span>".format(color, v))


def ai_box(content):
    st.markdown(
        "<div class='ai-box'>"
        "<div class='ai-box-title'>AI Insight</div>"
        "{}</div>".format(content),
        unsafe_allow_html=True)


def resolve_url(row):
    url = str(row.get("Hyperlink", "")).strip()
    if not url or url == "nan":
        url = str(row.get("File Link", "")).strip()
    if not url or url == "nan":
        url = str(row.get("File URL",  "")).strip()
    return "" if url == "nan" else url


def kpi(col, val, lbl, bg):
    col.markdown(
        "<div class='kpi-box' style='background:{}'>"
        "<div class='kpi-value'>{}</div>"
        "<div class='kpi-label'>{}</div>"
        "</div>".format(bg, val, lbl),
        unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# SIDEBAR (only shown when data is available)
# ════════════════════════════════════════════════════════════
selected_svcs  = []
selected_cat   = "All"
selected_vendor = "All"
d_filt         = pd.DataFrame()

if not NO_DATA:
    with st.sidebar:
        st.markdown(
            "<div style='text-align:center;padding:20px 0 14px'>"
            "<div style='font-size:2em'>📋</div>"
            "<div style='font-size:1.05em;font-weight:700;"
            "color:white;margin:5px 0 2px'>IT Procurement</div>"
            "<div style='font-size:0.70em;color:#aaa;"
            "letter-spacing:1px;text-transform:uppercase'>"
            "Intelligence Dashboard</div></div>"
            "<hr style='border-color:#D04A02;"
            "border-width:2px;margin:0 0 14px'>",
            unsafe_allow_html=True)

        if DATA_SOURCE == "uploaded":
            st.markdown(
                "<div style='background:#D04A02;color:white;"
                "padding:6px 10px;border-radius:2px;"
                "font-size:0.75em;font-weight:700;"
                "text-align:center;margin-bottom:10px'>"
                "📤 USING UPLOADED CATALOG</div>",
                unsafe_allow_html=True)

        sb_label("📂 Category")
        all_cats = ["All"] + sorted([
            c for c in df_master["Category"].unique()
            if str(c).strip() not in ["", "nan"]
        ])
        selected_cat = st.selectbox(
            "Category", all_cats,
            label_visibility="collapsed")

        sb_label("🏢 Vendor")
        vpool = (
            df_master if selected_cat == "All"
            else df_master[
                df_master["Category"] == selected_cat])
        all_vendors = ["All"] + sorted([
            v for v in vpool["Vendor"].unique()
            if str(v).strip() not in ["", "nan"]
        ])
        selected_vendor = st.selectbox(
            "Vendor", all_vendors,
            label_visibility="collapsed")

        st.markdown(
            "<hr style='border-color:#555;margin:12px 0'>",
            unsafe_allow_html=True)

        d_filt = df_exploded.copy()
        if selected_cat    != "All":
            d_filt = d_filt[
                d_filt["Category"] == selected_cat]
        if selected_vendor != "All":
            d_filt = d_filt[
                d_filt["Vendor"]   == selected_vendor]

        sb_label("🔍 Search Services")
        svc_search = st.text_input(
            "Search", placeholder="e.g. Cisco, Oracle…",
            label_visibility="collapsed")

        all_svcs = sorted([
            s for s in d_filt["Service"].unique()
            if str(s).strip() not in ["", "nan"]
        ])
        if svc_search:
            all_svcs = [
                s for s in all_svcs
                if svc_search.lower() in s.lower()
            ]

        svc_groups = group_services(all_svcs)

        sb_label("🛠 Select Services by Group")
        st.markdown(
            "<p style='color:#aaa;font-size:0.75em;"
            "margin:0 0 8px'>Expand a group to select</p>",
            unsafe_allow_html=True)

        for group_name, group_svcs in svc_groups.items():
            with st.expander(
                "{} ({})".format(
                    group_name, len(group_svcs)),
                expanded=False,
            ):
                col_a, col_b = st.columns(2)
                if col_a.button(
                    "All",
                    key="all_{}".format(group_name),
                    use_container_width=True,
                ):
                    st.session_state[
                        "sel_{}".format(group_name)
                    ] = group_svcs
                if col_b.button(
                    "Clear",
                    key="clr_{}".format(group_name),
                    use_container_width=True,
                ):
                    st.session_state[
                        "sel_{}".format(group_name)
                    ] = []

                sk      = "sel_{}".format(group_name)
                default = st.session_state.get(sk, [])
                chosen  = st.multiselect(
                    "Pick",
                    options=group_svcs,
                    default=default,
                    label_visibility="collapsed",
                    key="ms_{}".format(group_name),
                )
                st.session_state[sk] = chosen
                selected_svcs.extend(chosen)

        st.markdown(
            "<hr style='border-color:#555;margin:12px 0'>",
            unsafe_allow_html=True)
        st.markdown(
            "<p style='color:#888;font-size:0.78em;"
            "margin:2px 0'>"
            "📄 {} quotes | 🛠 {} services | "
            "🏢 {} vendors</p>".format(
                len(df_master),
                df_exploded["Service"].nunique(),
                df_master["Vendor"].nunique()),
            unsafe_allow_html=True)

        if selected_svcs:
            st.markdown(
                "<p style='color:#D04A02;font-size:0.80em;"
                "font-weight:700;margin:4px 0'>"
                "✅ {} service(s) selected</p>".format(
                    len(selected_svcs)),
                unsafe_allow_html=True)

        # Reset catalog button
        if DATA_SOURCE == "uploaded":
            st.markdown(
                "<hr style='border-color:#555;margin:12px 0'>",
                unsafe_allow_html=True)
            if st.button(
                "🔄 Reset to Default Catalog",
                use_container_width=True,
            ):
                st.session_state[
                    "uploaded_catalog_df"] = None
                st.session_state[
                    "uploaded_catalog_exp"] = None
                st.rerun()


# ════════════════════════════════════════════════════════════
# MAIN HEADER
# ════════════════════════════════════════════════════════════
st.markdown(
    "<div style='background:#2D2D2D;color:white;"
    "padding:20px 28px;border-radius:4px;"
    "border-left:6px solid #D04A02;margin-bottom:22px'>"
    "<div style='font-size:0.72em;font-weight:700;"
    "letter-spacing:2px;text-transform:uppercase;"
    "color:#D04A02;margin-bottom:5px'>"
    "IT Procurement Analytics</div>"
    "<h1 style='margin:0;font-size:1.45em;font-weight:700;"
    "color:white'>Procurement Intelligence Dashboard</h1>"
    "<p style='margin:6px 0 0;opacity:0.6;font-size:0.85em'>"
    "Browse quotations · Compare prices · "
    "Upload &amp; score · AI insights · Catalog upload"
    "</p></div>",
    unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# KPI ROW
# ════════════════════════════════════════════════════════════
if not NO_DATA:
    k1, k2, k3, k4 = st.columns(4)
    kpi(k1, d_filt["File Name"].nunique() if not d_filt.empty
        else df_master["File Name"].nunique(),
        "Total Quotes",    "#D04A02")
    kpi(k2, d_filt["Service"].nunique() if not d_filt.empty
        else df_exploded["Service"].nunique(),
        "Unique Services", "#295477")
    kpi(k3, d_filt["Vendor"].nunique() if not d_filt.empty
        else df_master["Vendor"].nunique(),
        "Vendors",         "#299D8F")
    kpi(k4, d_filt["Category"].nunique() if not d_filt.empty
        else df_master["Category"].nunique(),
        "Categories",      "#2D2D2D")
    st.markdown("<br>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# TABS
# ════════════════════════════════════════════════════════════
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Analytics",
    "📋 Browse & Compare",
    "📤 Upload & Score",
    "📄 Data Table",
    "🗂 Upload Catalog",
])


# ════════════════════════════════════════════════════════════
# TAB 1 — ANALYTICS
# ════════════════════════════════════════════════════════════
with tab1:
    if NO_DATA:
        st.info(
            "No catalog loaded. "
            "Go to the **🗂 Upload Catalog** tab to get started.")
    else:
        use_df = d_filt if not d_filt.empty else df_exploded

        col_l, col_r = st.columns(2, gap="large")

        with col_l:
            section_title(
                "SERVICE OVERLAP ANALYSIS",
                "Orange = service quoted by multiple vendors.")
            shared = (
                use_df.groupby("Service")["Vendor"].nunique()
                .sort_values(ascending=False)
                .head(20).reset_index()
            )
            shared.columns = ["Service", "Vendor Count"]
            shared["Color"] = shared["Vendor Count"].apply(
                lambda x: "#D04A02" if x > 1 else "#C0C0C0")
            fig1 = go.Figure(go.Bar(
                x=shared["Vendor Count"],
                y=shared["Service"].str[:44],
                orientation="h",
                marker_color=shared["Color"],
                marker_line_width=0,
                text=shared["Vendor Count"],
                textposition="outside",
                textfont=dict(size=10),
            ))
            fig1.update_layout(
                height=480, plot_bgcolor=CBG,
                paper_bgcolor=CBG,
                margin=dict(l=5, r=40, t=20, b=10),
                font=CFONT,
                xaxis=dict(
                    title="Vendors", showgrid=True,
                    gridcolor="#E0E0E0", zeroline=False),
                yaxis=dict(
                    autorange="reversed",
                    tickfont=dict(size=9.5)),
                bargap=0.35)
            st.plotly_chart(fig1, use_container_width=True)

        with col_r:
            section_title(
                "VENDOR SERVICE COVERAGE",
                "Higher = broader vendor capability.")
            spv = (
                use_df.groupby("Vendor")["Service"].nunique()
                .sort_values(ascending=False).reset_index()
            )
            spv.columns = ["Vendor", "Count"]
            spv["Color"] = [
                vendor_color_map.get(v, "#8C8C8C")
                for v in spv["Vendor"]
            ]
            fig2 = go.Figure(go.Bar(
                x=spv["Vendor"], y=spv["Count"],
                marker_color=spv["Color"],
                marker_line_width=0,
                text=spv["Count"],
                textposition="outside",
                textfont=dict(size=10),
            ))
            fig2.update_layout(
                height=480, plot_bgcolor=CBG,
                paper_bgcolor=CBG,
                margin=dict(l=5, r=10, t=20, b=10),
                font=CFONT,
                yaxis=dict(
                    title="Unique Services",
                    showgrid=True,
                    gridcolor="#E0E0E0",
                    zeroline=False),
                xaxis=dict(
                    tickangle=-35,
                    tickfont=dict(size=9.5)),
                bargap=0.35)
            st.plotly_chart(fig2, use_container_width=True)

        section_title(
            "PROCUREMENT CATEGORY DISTRIBUTION",
            "Share of quote files across categories.")
        cat_c = (
            use_df.drop_duplicates(
                subset=["Category", "File Name"])
            .groupby("Category").size().reset_index()
        )
        cat_c.columns = ["Category", "Count"]
        if not cat_c.empty:
            fig3 = px.pie(
                cat_c, names="Category", values="Count",
                hole=0.50,
                color_discrete_sequence=COLORS)
            fig3.update_traces(
                textposition="outside",
                textinfo="label+percent",
                textfont_size=11,
                pull=[0.03] * len(cat_c))
            fig3.update_layout(
                height=380,
                margin=dict(l=20, r=20, t=20, b=20),
                paper_bgcolor=CBG, font=CFONT,
                legend=dict(
                    orientation="v", x=1.02, y=0.5,
                    font=dict(size=10)))
            st.plotly_chart(fig3, use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)
        section_title("AI VENDOR SUMMARY")
        svc_by_vendor = {}
        for v in df_master["Vendor"].unique():
            svcs = list(df_exploded[
                df_exploded["Vendor"] == v
            ]["Service"].unique())
            svc_by_vendor[v] = svcs
        ai_box(ai_service_summary(svc_by_vendor))


# ════════════════════════════════════════════════════════════
# TAB 2 — BROWSE & COMPARE
# ════════════════════════════════════════════════════════════
with tab2:
    if NO_DATA:
        st.info(
            "No catalog loaded. "
            "Go to **🗂 Upload Catalog** tab first.")
    elif not selected_svcs:
        st.info(
            "👈 Select services from the sidebar groups "
            "to browse matching quotations.")
    else:
        use_filt = d_filt if not d_filt.empty else df_exploded
        d_sel    = use_filt[
            use_filt["Service"].isin(selected_svcs)
        ].copy()

        if d_sel.empty:
            st.warning("No results found.")
        else:
            vsmap = defaultdict(set)
            for _, r in d_sel.iterrows():
                vsmap[r["Vendor"]].add(r["Service"])

            vendors_all  = sorted([
                v for v, s in vsmap.items()
                if set(selected_svcs).issubset(s)])
            vendors_some = sorted([
                v for v, s in vsmap.items()
                if not set(selected_svcs).issubset(s)])

            if len(selected_svcs) > 1:
                if vendors_all:
                    names = " · ".join(
                        ["**{}**".format(v)
                         for v in vendors_all])
                    st.success(
                        "✅ {} vendor(s) cover ALL {} "
                        "services: {}".format(
                            len(vendors_all),
                            len(selected_svcs), names))
                else:
                    st.warning(
                        "No single vendor covers all {} "
                        "services.".format(
                            len(selected_svcs)))
                if vendors_some:
                    with st.expander(
                            "Vendors with partial coverage",
                            expanded=False):
                        for v in vendors_some:
                            cov = vsmap[v].intersection(
                                set(selected_svcs))
                            c   = vendor_color_map.get(
                                v, "#8C8C8C")
                            st.markdown(
                                "{} covers **{}/{}**: "
                                "_{}_".format(
                                    vendor_pill(v, c),
                                    len(cov),
                                    len(selected_svcs),
                                    ", ".join(sorted(cov))),
                                unsafe_allow_html=True)

            section_title(
                "QUOTATION FILES — PER SERVICE",
                "Price Score: 100=cheapest, 0=most expensive.")

            has_price = "Quoted Price" in d_sel.columns

            for svc in selected_svcs:
                d_svc = (
                    d_sel[d_sel["Service"] == svc]
                    .drop_duplicates(
                        subset=["Vendor", "File Name"])
                    .sort_values("Vendor")
                )
                vc    = d_svc["Vendor"].nunique()
                s_tag = ("SHARED" if vc > 1
                          else "SINGLE VENDOR")

                with st.expander(
                    "{}  —  {} vendor(s) · {} file(s) · {}"
                    .format(svc, vc, len(d_svc), s_tag),
                    expanded=True,
                ):
                    pills = " ".join([
                        vendor_pill(
                            v,
                            vendor_color_map.get(
                                v, "#8C8C8C"))
                        for v in sorted(
                            d_svc["Vendor"].unique())
                    ])
                    st.markdown(
                        "<div style='margin-bottom:12px'>"
                        "<b style='font-size:0.87em'>"
                        "Vendors:</b>&nbsp;&nbsp;"
                        "{}</div>".format(pills),
                        unsafe_allow_html=True)

                    all_prices_svc = []
                    for _, r in d_svc.iterrows():
                        qp = _parse_num(
                            str(r.get("Quoted Price",
                                      "")).strip())
                        if qp > 0:
                            all_prices_svc.append(qp)
                        ck = "px_{}".format(
                            str(r.get(
                                "File Name","")).strip())
                        ca = st.session_state.get(ck)
                        if ca and ca.get("price_num",0) > 0:
                            all_prices_svc.append(
                                ca["price_num"])

                    rows = [
                        "<table class='comp-table'>"
                        "<thead><tr>"
                        "<th>Vendor</th>"
                        "<th>Category</th>"
                        "<th>File Name</th>"
                    ]
                    if has_price:
                        rows.append("<th>Quoted Price</th>")
                    rows.append(
                        "<th>Extracted Price</th>"
                        "<th>Price Score</th>"
                        "<th>Verdict</th>"
                        "<th>Open</th>"
                        "</tr></thead><tbody>")

                    vendor_prices_map = {}

                    for i, (_, row) in enumerate(
                            d_svc.iterrows()):
                        bg    = ("#ffffff" if i % 2 == 0
                                 else "#F3F3F3")
                        color = vendor_color_map.get(
                            row["Vendor"], "#8C8C8C")
                        fname = str(
                            row.get("File Name","")).strip()
                        url   = resolve_url(row)

                        v_cell = vendor_pill(
                            row["Vendor"], color)
                        fn_cell = (
                            "<span style='font-family:"
                            "monospace;font-size:0.79em;"
                            "word-break:break-all'>"
                            "{}</span>".format(fname))
                        l_cell = (
                            "<a href='{}' target='_blank' "
                            "style='color:#D04A02;"
                            "font-weight:600;"
                            "text-decoration:none'>"
                            "Open</a>".format(url)
                            if url and url.startswith("http")
                            else
                            "<span style='color:#bbb'>—</span>"
                        )

                        qp_str = str(
                            row.get("Quoted Price","")).strip()
                        qp_num = _parse_num(qp_str)
                        qp_cell = (
                            "<span style='color:#22992E;"
                            "font-weight:700;"
                            "font-family:monospace'>"
                            "{}</span>".format(_fmt(qp_str))
                            if qp_num > 0
                            else
                            "<span style='color:#bbb'>—</span>"
                        )

                        ck     = "px_{}".format(fname)
                        cached = st.session_state.get(ck)
                        if cached and cached.get(
                                "price_num", 0) > 0:
                            ep_num    = cached["price_num"]
                            ep_cell   = (
                                "<span style='color:#295477;"
                                "font-weight:700;"
                                "font-family:monospace'>"
                                "{}</span>".format(
                                    _fmt(cached["price"])))
                            ref_price = ep_num
                        elif qp_num > 0:
                            ep_cell   = (
                                "<span style='color:#bbb;"
                                "font-size:0.80em'>—</span>")
                            ref_price = qp_num
                        else:
                            ep_cell   = (
                                "<span style='color:#bbb;"
                                "font-size:0.80em'>—</span>")
                            ref_price = 0

                        vendor_prices_map[
                            row["Vendor"]] = ref_price

                        others = [
                            p for p in all_prices_svc
                            if p != ref_price
                        ]
                        if ref_price > 0 and others:
                            ps, _  = price_score(
                                ref_price, others)
                            sc     = score_color(ps)
                            ps_cell = (
                                "<span style='font-weight:"
                                "800;font-size:1.0em;"
                                "color:{}'>{}/100</span>"
                                .format(
                                    sc,
                                    ps if ps is not None
                                    else "—"))
                            if ps is not None:
                                if ps >= 70:
                                    verdict = (
                                        "<span style='color:"
                                        "#22992E;"
                                        "font-weight:700'>"
                                        "Competitive</span>")
                                elif ps >= 40:
                                    verdict = (
                                        "<span style='color:"
                                        "#FFB600;"
                                        "font-weight:700'>"
                                        "Average</span>")
                                else:
                                    verdict = (
                                        "<span style='color:"
                                        "#E0301E;"
                                        "font-weight:700'>"
                                        "High</span>")
                            else:
                                verdict = (
                                    "<span style='color:"
                                    "#bbb'>—</span>")
                        else:
                            ps_cell = (
                                "<span style='color:#bbb'>"
                                "—</span>")
                            verdict = (
                                "<span style='color:#bbb'>"
                                "—</span>")

                        rows.append(
                            "<tr style='background:{}'>"
                            "<td>{}</td>"
                            "<td style='color:#555'>{}</td>"
                            "<td>{}</td>".format(
                                bg, v_cell,
                                row["Category"], fn_cell))
                        if has_price:
                            rows.append(
                                "<td>{}</td>".format(qp_cell))
                        rows.append(
                            "<td>{}</td>"
                            "<td style='text-align:center'>"
                            "{}</td>"
                            "<td>{}</td>"
                            "<td>{}</td></tr>".format(
                                ep_cell, ps_cell,
                                verdict, l_cell))

                    rows.append("</tbody></table>")
                    st.markdown(
                        "".join(rows),
                        unsafe_allow_html=True)

                    all_p = [p for p in all_prices_svc
                             if p > 0]
                    if all_p:
                        st.markdown(
                            "<br>", unsafe_allow_html=True)
                        ai_box(ai_price_insight(
                            0, all_p,
                            {v: p for v, p
                             in vendor_prices_map.items()
                             if p > 0}))

                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button(
                        "Extract Prices — {}".format(
                            svc[:40]),
                        key="ep_{}".format(svc[:35]),
                        type="primary",
                    ):
                        prog = st.progress(0)
                        n    = len(d_svc)
                        for ki, (_, row) in enumerate(
                                d_svc.iterrows()):
                            fname = str(
                                row.get(
                                    "File Name","")).strip()
                            url   = resolve_url(row)
                            ck    = "px_{}".format(fname)
                            if (url and
                                    url.startswith("http") and
                                    st.session_state.get(
                                        ck) is None):
                                res = extract_price_from_url(
                                    url)
                                st.session_state[ck] = res
                            prog.progress((ki + 1) / n)
                        prog.empty()
                        st.rerun()

                    price_data = []
                    for _, row in d_svc.iterrows():
                        fname  = str(
                            row.get("File Name","")).strip()
                        ck     = "px_{}".format(fname)
                        cached = st.session_state.get(ck)
                        qp     = _parse_num(str(
                            row.get(
                                "Quoted Price","")).strip())
                        ep     = (cached["price_num"]
                                  if cached else 0.0)
                        pval   = ep if ep > 0 else qp
                        if pval > 0:
                            price_data.append({
                                "Label": "{}\n{}".format(
                                    row["Vendor"],
                                    fname[:18]),
                                "Price": pval,
                                "Color": vendor_color_map.get(
                                    row["Vendor"], "#8C8C8C"),
                            })

                    if len(price_data) >= 2:
                        st.markdown(
                            "<br>", unsafe_allow_html=True)
                        section_title(
                            "PRICE COMPARISON — {}".format(
                                svc[:50]))
                        pdf2 = pd.DataFrame(price_data)
                        avg_p2 = (
                            sum(r["Price"]
                                for r in price_data) /
                            len(price_data))
                        mf = go.Figure(go.Bar(
                            x=pdf2["Label"],
                            y=pdf2["Price"],
                            marker_color=pdf2["Color"],
                            marker_line_width=0,
                            text=pdf2["Price"].apply(_fmt),
                            textposition="outside",
                        ))
                        mf.add_hline(
                            y=avg_p2,
                            line_dash="dash",
                            line_color="#FFB600",
                            line_width=2,
                            annotation_text="Avg: {}".format(
                                _fmt(avg_p2)),
                            annotation_position="top right")
                        mf.update_layout(
                            height=300,
                            plot_bgcolor=CBG,
                            paper_bgcolor=CBG,
                            margin=dict(
                                l=5, r=10, t=20, b=10),
                            font=CFONT,
                            yaxis=dict(
                                title="Price",
                                showgrid=True,
                                gridcolor="#E0E0E0",
                                zeroline=False),
                            xaxis=dict(tickangle=-20),
                            bargap=0.4)
                        st.plotly_chart(
                            mf, use_container_width=True)


# ════════════════════════════════════════════════════════════
# TAB 3 — UPLOAD & SCORE
# ════════════════════════════════════════════════════════════
with tab3:
    if NO_DATA:
        st.info(
            "No catalog loaded. "
            "Go to **🗂 Upload Catalog** tab first.")
    else:
        st.markdown(
            "<div style='background:#2D2D2D;color:white;"
            "padding:14px 20px;border-radius:4px;"
            "border-left:6px solid #D04A02;"
            "margin-bottom:16px'>"
            "<div style='font-size:0.72em;font-weight:700;"
            "letter-spacing:2px;text-transform:uppercase;"
            "color:#D04A02;margin-bottom:4px'>"
            "New Quotation Analysis</div>"
            "<div style='font-size:1.0em;font-weight:700'>"
            "Upload a new quotation → auto-extract price → "
            "score vs history</div>"
            "<div style='font-size:0.83em;opacity:0.6;"
            "margin-top:4px'>"
            "Supports PDF, XLSX, DOCX · "
            "Price Score 0-100 · Similarity matching"
            "</div></div>",
            unsafe_allow_html=True)

        section_title("STEP 1 — UPLOAD FILE")
        uploaded = st.file_uploader(
            "Upload",
            type=["pdf", "xlsx", "xls", "docx"],
            label_visibility="collapsed",
            key="quot_upload")

        if uploaded is not None:
            content   = uploaded.read()
            ext       = uploaded.name.rsplit(".", 1)[-1].lower()
            fname_up  = uploaded.name
            st.success("Uploaded: **{}** ({} KB)".format(
                fname_up, round(len(content)/1024, 1)))

            section_title("STEP 2 — EXTRACTED PRICE")
            with st.spinner("Extracting price…"):
                result    = extract_price_from_bytes(
                    content, ext)
                new_price = result["price_num"]
                new_text  = result["text"]

            if new_price > 0:
                st.markdown(
                    "<div class='score-card score-low'>"
                    "<div style='font-size:0.72em;"
                    "font-weight:700;letter-spacing:1px;"
                    "text-transform:uppercase;"
                    "color:#22992E'>Extracted Price</div>"
                    "<div class='score-num' "
                    "style='color:#22992E'>{}</div>"
                    "</div>".format(_fmt(new_price)),
                    unsafe_allow_html=True)
            else:
                st.warning(
                    "Could not extract price automatically.")
                manual = st.number_input(
                    "Enter price manually (USD)",
                    min_value=0.0, step=100.0, value=0.0,
                    key="man_price")
                if manual > 0:
                    new_price = manual

            section_title(
                "STEP 3 — SELECT SERVICES")
            all_svcs_up = sorted([
                s for s in df_exploded["Service"].unique()
                if str(s).strip() not in ["", "nan"]
            ])
            svc_search_up = st.text_input(
                "Search",
                placeholder="Filter services…",
                key="svc_up2",
                label_visibility="collapsed")
            if svc_search_up:
                all_svcs_up = [
                    s for s in all_svcs_up
                    if svc_search_up.lower() in s.lower()
                ]
            svc_groups_up = group_services(all_svcs_up)
            new_services  = []
            for gname, gsvcs in svc_groups_up.items():
                with st.expander(
                    "{} ({})".format(gname, len(gsvcs)),
                    expanded=False
                ):
                    chosen_up = st.multiselect(
                        "Pick", options=gsvcs, default=[],
                        label_visibility="collapsed",
                        key="up_ms2_{}".format(gname))
                    new_services.extend(chosen_up)

            cat_filter_up = st.selectbox(
                "Filter historical by category",
                options=["All"] + sorted([
                    c for c in df_master["Category"].unique()
                    if str(c).strip() not in ["", "nan"]]),
                key="cat_up2")

            section_title("STEP 4 — COMPARISON RESULTS")

            if not new_services and new_price <= 0:
                st.info(
                    "Select services and/or provide a price.")
            else:
                if new_services:
                    candidates = df_exploded[
                        df_exploded["Service"].isin(
                            new_services)].copy()
                else:
                    candidates = df_exploded.copy()

                if cat_filter_up != "All":
                    candidates = candidates[
                        candidates["Category"] ==
                        cat_filter_up]

                cand_files = (
                    candidates
                    .drop_duplicates(
                        subset=["File Name", "Vendor"])
                    [["File Name","Vendor","Category",
                      "Hyperlink","Quoted Price"]]
                    .copy()
                )

                if cand_files.empty:
                    st.warning(
                        "No historical quotes found.")
                else:
                    hist_prices_list = []
                    vendor_p_map     = {}
                    for _, r in cand_files.iterrows():
                        qp = _parse_num(str(
                            r.get("Quoted Price","")).strip())
                        if qp > 0:
                            hist_prices_list.append(qp)
                            vendor_p_map[r["Vendor"]] = qp
                        ck = "px_{}".format(str(
                            r.get("File Name","")).strip())
                        ca = st.session_state.get(ck)
                        if ca and ca.get("price_num",0) > 0:
                            hist_prices_list.append(
                                ca["price_num"])
                            vendor_p_map[r["Vendor"]] = \
                                ca["price_num"]

                    if new_price > 0 and hist_prices_list:
                        ps, ps_label = price_score(
                            new_price, hist_prices_list)
                        avg_h = (sum(hist_prices_list) /
                                 len(hist_prices_list))
                        mn_h  = min(hist_prices_list)
                        mx_h  = max(hist_prices_list)

                        sc1, sc2, sc3, sc4 = st.columns(4)
                        sc1.markdown(
                            "<div class='score-card {}'>"
                            "<div style='font-size:0.70em;"
                            "font-weight:700;"
                            "letter-spacing:1px;"
                            "text-transform:uppercase;"
                            "color:{}'>Price Score</div>"
                            "<div class='score-num' "
                            "style='color:{}'>{}/100</div>"
                            "<div style='font-size:0.75em;"
                            "color:#555;margin-top:4px'>"
                            "vs {} quotes</div>"
                            "</div>".format(
                                score_css(ps),
                                score_color(ps),
                                score_color(ps),
                                ps if ps is not None
                                else "N/A",
                                len(hist_prices_list)),
                            unsafe_allow_html=True)

                        sc2.markdown(
                            "<div class='score-card "
                            "score-medium'>"
                            "<div style='font-size:0.70em;"
                            "font-weight:700;"
                            "letter-spacing:1px;"
                            "text-transform:uppercase;"
                            "color:#FFB600'>Your Price</div>"
                            "<div class='score-num' "
                            "style='color:#D04A02'>"
                            "{}</div></div>".format(
                                _fmt(new_price)),
                            unsafe_allow_html=True)

                        sc3.markdown(
                            "<div class='score-card "
                            "score-medium'>"
                            "<div style='font-size:0.70em;"
                            "font-weight:700;"
                            "letter-spacing:1px;"
                            "text-transform:uppercase;"
                            "color:#FFB600'>"
                            "Hist. Average</div>"
                            "<div class='score-num' "
                            "style='color:#295477'>"
                            "{}</div>"
                            "<div style='font-size:0.72em;"
                            "color:#555;margin-top:4px'>"
                            "min {} · max {}</div>"
                            "</div>".format(
                                _fmt(avg_h),
                                _fmt(mn_h), _fmt(mx_h)),
                            unsafe_allow_html=True)

                        sc4.markdown(
                            "<div class='score-card {}'>"
                            "<div style='font-size:0.70em;"
                            "font-weight:700;"
                            "letter-spacing:1px;"
                            "text-transform:uppercase;"
                            "color:{}'>Verdict</div>"
                            "<div style='font-size:0.88em;"
                            "font-weight:700;color:{};"
                            "margin-top:6px'>{}</div>"
                            "</div>".format(
                                score_css(ps),
                                score_color(ps),
                                score_color(ps), ps_label),
                            unsafe_allow_html=True)

                        st.markdown(
                            "<br>", unsafe_allow_html=True)
                        ai_box(ai_price_insight(
                            new_price, hist_prices_list,
                            vendor_p_map))

                        # Price chart
                        st.markdown(
                            "<br>", unsafe_allow_html=True)
                        section_title("PRICE POSITIONING")
                        chart_data = []
                        for _, r in cand_files.iterrows():
                            fn     = str(r.get(
                                "File Name","")).strip()
                            ck     = "px_{}".format(fn)
                            cached = st.session_state.get(ck)
                            qp     = _parse_num(str(r.get(
                                "Quoted Price","")).strip())
                            ep     = (cached["price_num"]
                                      if cached else 0.0)
                            pval   = ep if ep > 0 else qp
                            if pval > 0:
                                chart_data.append({
                                    "Label": "{}/{}".format(
                                        r["Vendor"],
                                        fn[:15]),
                                    "Price": pval,
                                    "Type" : "Historical",
                                    "Color": vendor_color_map
                                    .get(r["Vendor"],
                                         "#8C8C8C"),
                                })
                        chart_data.append({
                            "Label": "NEW:{}".format(
                                fname_up[:15]),
                            "Price": new_price,
                            "Type" : "New Upload",
                            "Color": "#D04A02",
                        })
                        cdf = pd.DataFrame(
                            chart_data).sort_values("Price")
                        cf  = go.Figure()
                        cf.add_trace(go.Bar(
                            x=cdf[cdf["Type"]=="Historical"][
                                "Label"],
                            y=cdf[cdf["Type"]=="Historical"][
                                "Price"],
                            marker_color=cdf[
                                cdf["Type"]=="Historical"][
                                "Color"],
                            marker_line_width=0,
                            name="Historical",
                            text=cdf[
                                cdf["Type"]=="Historical"][
                                "Price"].apply(_fmt),
                            textposition="outside"))
                        cf.add_trace(go.Bar(
                            x=cdf[cdf["Type"]=="New Upload"][
                                "Label"],
                            y=cdf[cdf["Type"]=="New Upload"][
                                "Price"],
                            marker_color="#D04A02",
                            marker_line_width=0,
                            name="Your Upload",
                            text=cdf[
                                cdf["Type"]=="New Upload"][
                                "Price"].apply(_fmt),
                            textposition="outside"))
                        cf.add_hline(
                            y=avg_h,
                            line_dash="dash",
                            line_color="#FFB600",
                            line_width=2,
                            annotation_text="Avg:{}".format(
                                _fmt(avg_h)),
                            annotation_position="top right")
                        cf.update_layout(
                            height=360,
                            plot_bgcolor=CBG,
                            paper_bgcolor=CBG,
                            margin=dict(
                                l=5, r=10, t=20, b=10),
                            font=CFONT, barmode="group",
                            yaxis=dict(
                                title="Price",
                                showgrid=True,
                                gridcolor="#E0E0E0",
                                zeroline=False),
                            xaxis=dict(tickangle=-25),
                            legend=dict(
                                orientation="h",
                                x=0, y=1.05),
                            bargap=0.25)
                        st.plotly_chart(
                            cf, use_container_width=True)


# ════════════════════════════════════════════════════════════
# TAB 4 — DATA TABLE
# ════════════════════════════════════════════════════════════
with tab4:
    if NO_DATA:
        st.info(
            "No catalog loaded. "
            "Go to **🗂 Upload Catalog** tab first.")
    else:
        dm = df_master.copy()
        if selected_cat    != "All":
            dm = dm[dm["Category"] == selected_cat]
        if selected_vendor != "All":
            dm = dm[dm["Vendor"]   == selected_vendor]
        st.dataframe(
            dm.drop(
                columns=["Services List", "Hyperlink"],
                errors="ignore"),
            use_container_width=True,
            height=500)


# ════════════════════════════════════════════════════════════
# TAB 5 — UPLOAD CATALOG
# ════════════════════════════════════════════════════════════
with tab5:
    st.markdown(
        "<div style='background:#2D2D2D;color:white;"
        "padding:20px 28px;border-radius:4px;"
        "border-left:6px solid #D04A02;"
        "margin-bottom:24px'>"
        "<div style='font-size:0.72em;font-weight:700;"
        "letter-spacing:2px;text-transform:uppercase;"
        "color:#D04A02;margin-bottom:5px'>"
        "Catalog Management</div>"
        "<h1 style='margin:0;font-size:1.3em;"
        "font-weight:700;color:white'>"
        "Upload Master Catalog</h1>"
        "<p style='margin:6px 0 0;opacity:0.6;"
        "font-size:0.85em'>"
        "Upload any Excel or CSV quotation catalog — "
        "AI auto-detects columns, builds the dashboard "
        "and generates insights automatically."
        "</p></div>",
        unsafe_allow_html=True)

    # Current status
    if DATA_SOURCE == "uploaded":
        st.success(
            "✅ Currently using an uploaded catalog: "
            "**{}** rows, **{}** vendors, "
            "**{}** services".format(
                len(df_master),
                df_master["Vendor"].nunique(),
                df_exploded["Service"].nunique()))

    # Instructions
    with st.expander(
            "How to prepare your catalog file",
            expanded=False):
        st.markdown("""
**Required columns** (names detected automatically):

| Column | Examples of accepted names |
|---|---|
| **Category** | Category, Type, Domain, Product Type |
| **Vendor** | Vendor, Supplier, Company, Provider |
| **File Name** | File Name, Filename, Document, Attachment |
| **Services** | Comments, Services, Description, Scope |
| **Price** | Quoted Price, Price, Cost, Amount, Value |
| **Link** | File Link, URL, Hyperlink, Path |

**Supported formats:** `.xlsx`, `.xls`, `.csv`

**Tips:**
- Services/descriptions in the Comments column should be
  one service per line (line-break separated)
- File links should be full URLs starting with `http`
- The AI will auto-detect your header row
        """)

    st.markdown(
        "<div style='font-size:0.78em;font-weight:700;"
        "letter-spacing:1px;text-transform:uppercase;"
        "color:#D04A02;margin-bottom:12px'>"
        "UPLOAD YOUR CATALOG FILE</div>",
        unsafe_allow_html=True)

    catalog_file = st.file_uploader(
        "Upload Master Catalog",
        type=["xlsx", "xls", "csv"],
        label_visibility="collapsed",
        key="catalog_upload",
        help="Upload Excel (.xlsx/.xls) or CSV catalog file",
    )

    if catalog_file is not None:
        file_bytes = catalog_file.read()
        fname_cat  = catalog_file.name

        st.markdown(
            "<div class='catalog-step'>"
            "<div class='catalog-step-num'>"
            "Step 1 — File Received</div>"
            "<b>{}</b> — {} KB uploaded".format(
                fname_cat,
                round(len(file_bytes)/1024, 1))
            + "</div>",
            unsafe_allow_html=True)

        with st.spinner(
                "AI is analyzing your catalog…"):
            df_new, df_exp_new, err = \
                process_uploaded_catalog(
                    file_bytes, fname_cat)

        if err:
            st.error("❌ {}".format(err))
        elif df_new is None:
            st.error(
                "❌ Could not process file. "
                "Check format and try again.")
        else:
            # Step 2 — Column mapping
            st.markdown(
                "<div class='catalog-step'>"
                "<div class='catalog-step-num'>"
                "Step 2 — AI Column Detection</div>"
                "Successfully mapped columns. "
                "Found: <b>{}</b> rows across "
                "<b>{}</b> vendors and "
                "<b>{}</b> categories.".format(
                    len(df_new),
                    df_new["Vendor"].nunique(),
                    df_new["Category"].nunique())
                + "</div>",
                unsafe_allow_html=True)

            # Show detected columns
            st.markdown(
                "<div style='font-size:0.78em;"
                "font-weight:700;letter-spacing:1px;"
                "text-transform:uppercase;color:#D04A02;"
                "margin:12px 0 6px'>"
                "DETECTED COLUMNS</div>",
                unsafe_allow_html=True)

            col_disp = st.columns(
                min(len(df_new.columns), 4))
            std_cols = [
                "Category","Vendor","File Name",
                "Comments","Quoted Price",
                "File Link","Hyperlink","Services List"
            ]
            shown = 0
            for col in df_new.columns:
                if col in std_cols and shown < 4:
                    col_disp[shown % 4].markdown(
                        "<div style='background:#2D2D2D;"
                        "color:#D04A02;padding:8px 10px;"
                        "border-radius:2px;font-size:0.80em;"
                        "font-weight:700;text-align:center;"
                        "margin-bottom:4px'>"
                        "{}</div>"
                        "<div style='text-align:center;"
                        "font-size:0.75em;color:#555'>"
                        "{} unique values</div>".format(
                            col,
                            df_new[col].nunique()),
                        unsafe_allow_html=True)
                    shown += 1

            # Step 3 — AI Analysis
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown(
                "<div class='catalog-step'>"
                "<div class='catalog-step-num'>"
                "Step 3 — AI Analysis</div>"
                "Generating insights from catalog data…"
                "</div>",
                unsafe_allow_html=True)

            insights = ai_analyze_catalog(
                df_new, df_exp_new)

            # Display insights
            ins_cols = st.columns(2)
            insight_items = [
                ("Overview",
                 insights.get("overview", ""),
                 "#295477"),
                ("Top Vendor",
                 insights.get("top_vendor", ""),
                 "#299D8F"),
                ("Most Competitive Service",
                 insights.get("competitive", ""),
                 "#D04A02"),
                ("Category Concentration",
                 insights.get("category", ""),
                 "#FFB600"),
            ]
            if insights.get("pricing"):
                insight_items.append((
                    "Pricing Range",
                    insights["pricing"],
                    "#22992E"))

            for idx, (title, content, color) in \
                    enumerate(insight_items):
                if content:
                    ins_cols[idx % 2].markdown(
                        "<div class='insight-card'>"
                        "<div style='font-size:0.72em;"
                        "font-weight:700;letter-spacing:1px;"
                        "text-transform:uppercase;"
                        "color:{};margin-bottom:6px'>"
                        "{}</div>"
                        "<div style='font-size:0.88em;"
                        "color:#2D2D2D'>{}</div>"
                        "</div>".format(
                            color, title, content),
                        unsafe_allow_html=True)

            # Recommendations
            recs = insights.get("recommendations", [])
            if recs:
                st.markdown(
                    "<div style='font-size:0.78em;"
                    "font-weight:700;letter-spacing:1px;"
                    "text-transform:uppercase;"
                    "color:#D04A02;margin:16px 0 8px'>"
                    "AI RECOMMENDATIONS</div>",
                    unsafe_allow_html=True)
                for rec in recs:
                    st.markdown(
                        "<div style='background:#FFF3F0;"
                        "border-left:3px solid #D04A02;"
                        "padding:8px 12px;border-radius:2px;"
                        "margin-bottom:6px;"
                        "font-size:0.87em;color:#2D2D2D'>"
                        "▸ {}</div>".format(rec),
                        unsafe_allow_html=True)

            # Quick charts from new catalog
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown(
                "<div style='font-size:0.78em;"
                "font-weight:700;letter-spacing:1px;"
                "text-transform:uppercase;"
                "color:#D04A02;margin-bottom:8px'>"
                "CATALOG PREVIEW CHARTS</div>",
                unsafe_allow_html=True)

            pc1, pc2 = st.columns(2)

            with pc1:
                spv_new = (
                    df_exp_new.groupby("Vendor")[
                        "Service"].nunique()
                    .sort_values(ascending=False)
                    .reset_index()
                )
                spv_new.columns = ["Vendor", "Services"]
                vc_map_new = {
                    v: get_color(i)
                    for i, v in enumerate(
                        spv_new["Vendor"].tolist())
                }
                pf1 = go.Figure(go.Bar(
                    x=spv_new["Vendor"],
                    y=spv_new["Services"],
                    marker_color=[
                        vc_map_new.get(v, "#8C8C8C")
                        for v in spv_new["Vendor"]],
                    marker_line_width=0,
                    text=spv_new["Services"],
                    textposition="outside",
                ))
                pf1.update_layout(
                    title="Services per Vendor",
                    height=320,
                    plot_bgcolor=CBG,
                    paper_bgcolor=CBG,
                    margin=dict(l=5, r=10, t=40, b=10),
                    font=CFONT,
                    yaxis=dict(
                        showgrid=True,
                        gridcolor="#E0E0E0",
                        zeroline=False),
                    xaxis=dict(tickangle=-30),
                    bargap=0.35)
                st.plotly_chart(
                    pf1, use_container_width=True)

            with pc2:
                cat_new = (
                    df_new.drop_duplicates(
                        subset=["Category","File Name"])
                    .groupby("Category").size()
                    .reset_index()
                )
                cat_new.columns = ["Category", "Count"]
                if not cat_new.empty:
                    pf2 = px.pie(
                        cat_new,
                        names="Category",
                        values="Count",
                        hole=0.45,
                        color_discrete_sequence=COLORS)
                    pf2.update_traces(
                        textposition="outside",
                        textinfo="label+percent",
                        textfont_size=10)
                    pf2.update_layout(
                        title="Category Distribution",
                        height=320,
                        margin=dict(
                            l=10, r=10, t=40, b=10),
                        paper_bgcolor=CBG,
                        font=CFONT)
                    st.plotly_chart(
                        pf2, use_container_width=True)

            # Preview table
            st.markdown(
                "<div style='font-size:0.78em;"
                "font-weight:700;letter-spacing:1px;"
                "text-transform:uppercase;"
                "color:#D04A02;margin:16px 0 8px'>"
                "DATA PREVIEW (First 20 rows)</div>",
                unsafe_allow_html=True)
            st.dataframe(
                df_new.drop(
                    columns=["Services List",
                              "Hyperlink"],
                    errors="ignore").head(20),
                use_container_width=True,
                height=300)

            # Apply catalog button
            st.markdown("<br>", unsafe_allow_html=True)
            apply_col, _ = st.columns([2, 3])
            if apply_col.button(
                "✅ Apply This Catalog to Dashboard",
                type="primary",
                use_container_width=True,
                key="apply_catalog",
            ):
                st.session_state[
                    "uploaded_catalog_df"] = df_new
                st.session_state[
                    "uploaded_catalog_exp"] = df_exp_new
                st.success(
                    "Catalog applied! "
                    "The dashboard now uses your uploaded "
                    "catalog. Use the sidebar to filter "
                    "and explore.")
                st.rerun()
