# ============================================================
#  generate_demo_data.py
#  Run once to create:
#  - Master Catalog.xlsx  (with hyperlinks)
#  - demo_quotes/         (folder with sample quote files)
# ============================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ════════════════════════════════════════════════════════════
# DEMO QUOTATION DATA
# ════════════════════════════════════════════════════════════
QUOTES = [
    # NTT Data
    {
        "vendor"   : "NTT Data",
        "category" : "Network & Telecom",
        "filename" : "NTT_Cisco_Catalyst_C9300_Quote.xlsx",
        "services" : "Cisco Catalyst C9300\nCisco Catalyst C9200\nCisco Wireless 9176I",
        "price"    : 185000,
    },
    {
        "vendor"   : "NTT Data",
        "category" : "Network & Telecom",
        "filename" : "NTT_Network_Core_Infrastructure.xlsx",
        "services" : "Cisco Catalyst C8300\nCisco ISR 4451\nNetwork Installation",
        "price"    : 320000,
    },
    {
        "vendor"   : "NTT Data",
        "category" : "Cybersecurity",
        "filename" : "NTT_Firewall_Proposal.xlsx",
        "services" : "Palo Alto Firewall\nFortiGate 600E\nSecurity Audit",
        "price"    : 210000,
    },
    # TPS
    {
        "vendor"   : "TPS",
        "category" : "Network & Telecom",
        "filename" : "TPS_Cisco_Catalyst_Quote_2024.xlsx",
        "services" : "Cisco Catalyst C9300\nCisco Catalyst C9200L\nCisco Wireless 9172I",
        "price"    : 172000,
    },
    {
        "vendor"   : "TPS",
        "category" : "Network & Telecom",
        "filename" : "TPS_Wireless_Infra_Proposal.xlsx",
        "services" : "Cisco Wireless 9176I\nCisco Wireless 9172I\nWireless Controller",
        "price"    : 145000,
    },
    {
        "vendor"   : "TPS",
        "category" : "Hosting",
        "filename" : "TPS_Cloud_Hosting_Services.xlsx",
        "services" : "Azure Virtual Machines\nAzure Storage\nAzure Backup",
        "price"    : 95000,
    },
    # PCC Connection
    {
        "vendor"   : "PCC Connection",
        "category" : "Network & Telecom",
        "filename" : "PCC_Network_Equipment_2024.xlsx",
        "services" : "Cisco Catalyst C9300\nHPE ProCurve Switch\nNetwork Cabling",
        "price"    : 198000,
    },
    {
        "vendor"   : "PCC Connection",
        "category" : "End User Computing",
        "filename" : "PCC_Laptop_Desktop_Quote.xlsx",
        "services" : "Dell Latitude 5540\nHP EliteBook 840\nMicrosoft Surface Pro",
        "price"    : 280000,
    },
    # Colocation Environment Build
    {
        "vendor"   : "Colocation Env Build",
        "category" : "Hosting",
        "filename" : "CEB_DataCenter_Colocation.xlsx",
        "services" : "Data Centre Colocation\nRack Space Rental\nPower & Cooling",
        "price"    : 420000,
    },
    {
        "vendor"   : "Colocation Env Build",
        "category" : "Hosting",
        "filename" : "CEB_BOM_Japan_DC_Production.xlsx",
        "services" : "Azure Virtual Machines\nData Centre Colocation\nNetwork Connectivity",
        "price"    : 380000,
    },
    # SoftwareONE
    {
        "vendor"   : "SoftwareONE",
        "category" : "Software & Licensing",
        "filename" : "SONE_Microsoft365_Enterprise.xlsx",
        "services" : "Microsoft 365 E3\nMicrosoft 365 E5\nAzure Active Directory",
        "price"    : 520000,
    },
    {
        "vendor"   : "SoftwareONE",
        "category" : "Software & Licensing",
        "filename" : "SONE_Adobe_Oracle_Licensing.xlsx",
        "services" : "Adobe Creative Cloud\nOracle Database License\nVMware vSphere",
        "price"    : 340000,
    },
    # Dimension Data
    {
        "vendor"   : "Dimension Data",
        "category" : "Cybersecurity",
        "filename" : "DD_Security_SOC_Services.xlsx",
        "services" : "Palo Alto Firewall\nSplunk SIEM\nSOC Managed Services",
        "price"    : 290000,
    },
    {
        "vendor"   : "Dimension Data",
        "category" : "Network & Telecom",
        "filename" : "DD_Cisco_Catalyst_C9300_v2.xlsx",
        "services" : "Cisco Catalyst C9300\nCisco Nexus 9000\nData Centre Switching",
        "price"    : 195000,
    },
    {
        "vendor"   : "Dimension Data",
        "category" : "End User Computing",
        "filename" : "DD_EndUser_Devices_2024.xlsx",
        "services" : "Dell Latitude 5540\nLenovo ThinkPad X1\nApple MacBook Pro",
        "price"    : 310000,
    },
]

# Demo new quotation (uploaded by user for comparison)
NEW_QUOTE = {
    "vendor"   : "NEW VENDOR — Alpha Networks",
    "filename" : "DEMO_New_Quotation_AlphaNetworks.xlsx",
    "services" : ["Cisco Catalyst C9300",
                  "Cisco Wireless 9176I",
                  "Palo Alto Firewall"],
    "price"    : 203000,
    "items"    : [
        ("Cisco Catalyst C9300 48-port",  "unit",  120, 850),
        ("Cisco Wireless 9176I AP",        "unit",   80, 450),
        ("Palo Alto PA-3260 Firewall",     "unit",    4, 8500),
        ("Installation & Config",          "lump",    1, 15000),
        ("1 Year Support",                 "lump",    1, 12000),
    ],
}


# ════════════════════════════════════════════════════════════
# COLOUR HELPERS
# ════════════════════════════════════════════════════════════
def hex_fill(hex_color):
    return PatternFill(
        start_color=hex_color.replace("#",""),
        end_color=hex_color.replace("#",""),
        fill_type="solid")

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s,right=s,top=s,bottom=s)


# ════════════════════════════════════════════════════════════
# CREATE INDIVIDUAL QUOTE EXCEL FILES
# ════════════════════════════════════════════════════════════
def create_quote_file(path, q):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # Header band
    ws.merge_cells("A1:F1")
    ws["A1"] = "OFFICIAL QUOTATION"
    ws["A1"].font      = Font(bold=True, size=16,
                              color="FFFFFF")
    ws["A1"].fill      = hex_fill("#D04A02")
    ws["A1"].alignment = Alignment(
        horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Sub-header
    ws.merge_cells("A2:F2")
    ws["A2"] = q["vendor"]
    ws["A2"].font      = Font(bold=True, size=12,
                              color="FFFFFF")
    ws["A2"].fill      = hex_fill("#2D2D2D")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # Blank row
    ws.row_dimensions[3].height = 8

    # Quote info
    info = [
        ("Quotation Date", "15-Apr-2025"),
        ("Valid Until",    "15-Jul-2025"),
        ("Prepared For",   "Panasonic Automotive"),
        ("Prepared By",    q["vendor"]),
        ("Currency",       "USD"),
    ]
    for i, (lbl, val) in enumerate(info, start=4):
        ws["A{}".format(i)] = lbl
        ws["B{}".format(i)] = val
        ws["A{}".format(i)].font = Font(
            bold=True, size=10)
        ws["A{}".format(i)].fill = hex_fill("#F3F3F3")

    # Blank
    ws.row_dimensions[9].height = 8

    # Items header
    hdr_row = 10
    headers = ["#","Description","Unit",
               "Qty","Unit Price (USD)","Total (USD)"]
    hdr_fills = ["#D04A02","#D04A02","#D04A02",
                 "#D04A02","#D04A02","#D04A02"]
    for col, (h, f) in enumerate(
            zip(headers, hdr_fills), start=1):
        c = ws.cell(row=hdr_row, column=col)
        c.value     = h
        c.font      = Font(bold=True, size=10,
                           color="FFFFFF")
        c.fill      = hex_fill(f)
        c.alignment = Alignment(
            horizontal="center",
            vertical="center")
        c.border    = thin_border()
    ws.row_dimensions[hdr_row].height = 22

    # Determine line items
    if "items" in q:
        items = q["items"]
    else:
        svcs  = [s.strip() for s in
                 q["services"].split("\n") if s.strip()]
        base  = q["price"] // max(len(svcs), 1)
        items = [(svc,"lump",1,base) for svc in svcs]

    running = 0
    for j, item in enumerate(items, start=1):
        desc, unit, qty, up = item
        total    = qty * up
        running += total
        row = hdr_row + j
        bg  = "FFFFFF" if j % 2 else "F9F9F9"
        vals = [j, desc, unit, qty, up, total]
        for col, val in enumerate(vals, start=1):
            c        = ws.cell(row=row, column=col)
            c.value  = val
            c.fill   = hex_fill("#"+bg)
            c.border = thin_border()
            c.font   = Font(size=10)
            if col in (5, 6):
                c.number_format = "#,##0.00"
                c.alignment = Alignment(
                    horizontal="right")

    # Subtotal / Tax / Total rows
    last_item_row = hdr_row + len(items)
    tax   = round(running * 0.06, 2)
    grand = running + tax
    summary = [
        ("","","","","Subtotal",  running),
        ("","","","","Tax (6%)",  tax),
        ("","","","","GRAND TOTAL", grand),
    ]
    colors = ["F3F3F3","F3F3F3","D04A02"]
    ftcols = ["2D2D2D","2D2D2D","FFFFFF"]
    for k, (row_data, bg, fc) in enumerate(
            zip(summary, colors, ftcols)):
        row = last_item_row + k + 1
        for col in range(1, 7):
            c        = ws.cell(row=row, column=col)
            c.fill   = hex_fill("#"+bg)
            c.border = thin_border()
        ws.cell(row=row, column=5).value = row_data[4]
        ws.cell(row=row, column=5).font  = Font(
            bold=True, size=10, color=fc)
        ws.cell(row=row, column=5).alignment = Alignment(
            horizontal="right")
        ws.cell(row=row, column=6).value = row_data[5]
        ws.cell(row=row, column=6).font  = Font(
            bold=True, size=10, color=fc)
        ws.cell(row=row, column=6).number_format = "#,##0.00"
        ws.cell(row=row, column=6).alignment = Alignment(
            horizontal="right")

    # Footer note
    note_row = last_item_row + 6
    ws.merge_cells("A{}:F{}".format(note_row, note_row))
    ws["A{}".format(note_row)] = (
        "Note: All prices are in USD. "
        "This quotation is valid for 90 days. "
        "Prices subject to change without notice.")
    ws["A{}".format(note_row)].font = Font(
        italic=True, size=9, color="888888")

    # Column widths
    widths = [5, 40, 10, 8, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[
            get_column_letter(i)].width = w

    ws.sheet_view.showGridLines = False
    wb.save(path)


# ════════════════════════════════════════════════════════════
# CREATE MASTER CATALOG WITH HYPERLINKS
# ════════════════════════════════════════════════════════════
def create_master_catalog(quotes, demo_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Catalog"

    # Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = "IT PROCUREMENT — MASTER CATALOG"
    ws["A1"].font      = Font(bold=True, size=14,
                              color="FFFFFF")
    ws["A1"].fill      = hex_fill("#D04A02")
    ws["A1"].alignment = Alignment(
        horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Header
    headers = ["Category","Vendor","File Name",
               "Quoted Price","Comments","File Link"]
    for col, h in enumerate(headers, start=1):
        c        = ws.cell(row=2, column=col)
        c.value  = h
        c.font   = Font(bold=True, size=10,
                        color="FFFFFF")
        c.fill   = hex_fill("#2D2D2D")
        c.alignment = Alignment(
            horizontal="center")
        c.border = thin_border()
    ws.row_dimensions[2].height = 20

    # Data rows
    for i, q in enumerate(quotes, start=3):
        fpath = os.path.join(demo_dir, q["filename"])
        bg    = "FFFFFF" if i % 2 else "F3F3F3"

        vals = [
            q["category"],
            q["vendor"],
            q["filename"],
            q["price"],
            q["services"],
            fpath,
        ]
        for col, val in enumerate(vals, start=1):
            c        = ws.cell(row=i, column=col)
            c.value  = val
            c.fill   = hex_fill("#"+bg)
            c.border = thin_border()
            c.font   = Font(size=10)
            if col == 4:
                c.number_format = "#,##0.00"

        # Embed hyperlink in File Name cell
        fname_cell = ws.cell(row=i, column=3)
        fname_cell.hyperlink = fpath
        fname_cell.font = Font(
            size=10, color="D04A02",
            underline="single")

    # Column widths
    widths = [22, 24, 46, 16, 50, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[
            get_column_letter(i)].width = w

    ws.sheet_view.showGridLines = False
    ws.freeze_panes              = "A3"

    wb.save("Master Catalog.xlsx")
    print("✅ Master Catalog.xlsx created")


# ════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    demo_dir = "demo_quotes"
    os.makedirs(demo_dir, exist_ok=True)

    # Create individual quote files
    for q in QUOTES:
        path = os.path.join(demo_dir, q["filename"])
        create_quote_file(path, q)
        print("  Created: {}".format(q["filename"]))

    # Create new demo quotation
    new_path = os.path.join(
        demo_dir, NEW_QUOTE["filename"])
    create_quote_file(new_path, NEW_QUOTE)
    print("  Created: {}".format(NEW_QUOTE["filename"]))

    # Create master catalog
    create_master_catalog(QUOTES, demo_dir)

    print("\n✅ All demo files created successfully!")
    print("   Folder : demo_quotes/ ({} files)".format(
        len(QUOTES)+1))
    print("   Catalog: Master Catalog.xlsx")
    print("\n   Demo new quotation to upload:")
    print("   → demo_quotes/{}".format(
        NEW_QUOTE["filename"]))
